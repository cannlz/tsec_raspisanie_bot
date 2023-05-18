import gdown
import openpyxl
import re




async def downloadExel(link, namefile, foldername):
    # URL-адрес файла Excel на Google Drive
    url = ' '.join(re.findall(r'd/([^<>]+)/', link))
    newlink = f'https://drive.google.com/uc?id={url}'
    #print(link, namefile)
    # Путь для сохранения скачанного файла Excel
    output_path = f'{foldername}/{namefile}.xlsx'
    # Скачивание файла Excel с помощью gdown
    gdown.download(newlink, output_path, quiet=True, format='xlsx')
    return output_path


async def checktextRasp(groupselect, pathtofile):
    # Указываем путь к файлу Excel
    wb = openpyxl.load_workbook(pathtofile, data_only=True)

    # get first worksheet
    ws = wb.worksheets[0]

    # check first column in first 10 rows for fill color
    for row in range(1, 100):
        for column in range(1, 50):
            cell = ws.cell(column=column, row=row)
            bgColor = cell.fill.bgColor.index
            fgColor = cell.fill.fgColor.index
            if bgColor != '00000000':
                print(f"  bgColor={bgColor}")
            if fgColor != '00000000' and bgColor == 'FFFFCC99' or bgColor == 'FFFAC090':
                #print("good", cell.value)
                #print(f"  bgColor={bgColor}")
                #print(f"row {row}: {column}")
                try:
                    if cell.value == groupselect:
                        print("Найдена группа: ", cell.value)
                        listdata = []
                        rowset = row
                        for rows in range(1,100):
                            rowset = rowset + 1 
                            cell = ws.cell(column=column, row=rowset)
                            celltime = ws.cell(column=1, row=rowset)
                            bgColor = cell.fill.bgColor.index
                            #print("Задний фон след ячейки", bgColor)
                            if bgColor == '00000000' or bgColor == 0 or bgColor == 'FFFFFFFF':
                                if cell.value != None and cell.value != " ":
                                    
                                    #print(f"GOOD: {rowset}\n{cell.value}")
                                    #print(wb.sheetnames)
                                    if 'rasp_ites' in pathtofile:
                                        celltime = ws.cell(column=3, row=rowset)
                                        cellcount = ws.cell(column=2, row=rowset)
                                        listdata.append(f'№{cellcount.value} Предмет: {cell.value}\nВремя: {celltime.value}')
                                    elif 'rasp_rpco' in pathtofile:
                                        celltime = ws.cell(column=4, row=rowset)
                                        listdata.append(f'№{celltime.value} Предмет: {cell.value}')
                                    elif 'rasp_spsipb' in pathtofile:
                                        celltime = ws.cell(column=4, row=rowset)
                                        listdata.append(f'№{celltime.value} Предмет: {cell.value}')
                                    elif 'rasp_tc' in pathtofile:
                                        celltime = ws.cell(column=2, row=rowset)
                                        cellcount = ws.cell(column=1, row=rowset)
                                        listdata.append(f'№{cellcount.value} Предмет: {cell.value}\nВремя: {celltime.value}')
                                    elif 'rasp_ppkrc' in pathtofile:
                                        print("dwadwadawdawdwadawdaw: ", cell.value)
                                        celltime = ws.cell(column=2, row=rowset)
                                        cellcount = ws.cell(column=1, row=rowset)
                                        listdata.append(f'№{cellcount.value} Предмет: {cell.value}\nВремя: {celltime.value}')
                                    else:
                                        listdata.append(f'№{celltime.value} Предмет: {cell.value}')
                                    #print(f"bgColor={bgColor}\nfgColor={fgColor}")   
                            else:
                                #print(f"STOP: {rowset}\n{cell.value}")
                                #print(f"bgColor={bgColor}\nfgColor={fgColor}")
                                break
                        #print(listdata, wb.sheetnames)
                        return listdata, wb.sheetnames
                        
                except Exception as e:
                    print(e)
                             